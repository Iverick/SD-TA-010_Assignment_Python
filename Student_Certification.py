# SD-TA-010_Assignment/input - where input files are located
# SD-TA-010_Assignment/input - where output files located
# SD-TA-010_Assignment/Student_Certification.py - program script
# SD-TA-010_Assignment/Tasks/ - folder includes solution for the text tasks 1,2 and 4

import csv
import json
import os

from datetime import date

from openpyxl import Workbook

def main():
    assessmentResultsPath = "input_files/Assessment Results.csv"
    industryCertPath = "input_files/Industry Cert Results.csv"
    outputFolder = "output_files"
    '''
    assessmentResults supposed to have this structure:
    {studentName: {'grades': {courseName: 'PASS/FAIL'}, 'cloudCert': 'PASS', 'optionCert': 'PASS', 'overallScore': 'PASS'}, ...}
    '''
    assessmentResults = {}

    try:
        readAssessmentResults(assessmentResultsPath, assessmentResults)
        readIndustryCerts(industryCertPath, assessmentResults)
        calculateOverallScore(assessmentResults)
        # print(json.dumps(assessmentResults, indent=1))
        # For every student generate an output txt file
        for studentName in assessmentResults:
            generateXlslFile(assessmentResults[studentName], studentName, outputFolder)
        print("Generated all reports!")
    except FileNotFoundError as e:
        print(f"Error: input file not found: {e}")
    except Exception as e:
        print(f"Error in program execution: {e}")
    finally:
        print("Program finished!")


'''
Reads data from assessmentResults file with provided path and stores them into a assessmentResults dict

return updated assessmentResults
'''
def readAssessmentResults(filePath, assessmentResults):
    PASS_MARK = 50
    courseNames = []
    rows = []

    try:
        # Open an input file and transform each line into a list
        with open(filePath, 'r') as file:
            reader = csv.reader(file)
            for line in reader:
                rows.append(line)
    except FileNotFoundError:
        print(f"Error opening {filePath}")
        raise

    # Grab course names from the first row and push them into a list
    for i in range(len(rows[0])):
        if (i == 0):
            continue
        courseNames.append(rows[0][i])

    # I need to get student name and list of scores for every student
    for row in rows[2:]:
        studentGrades = {}
        # skip empty rows
        if not row[0]:
            continue

        studentName = row[0].strip()
        for i in range(len(courseNames)):
            # convert each grade into PASS or FAIL and store it to the corresponding courseName
            courseName = courseNames[i]
            grade = row [i + 1]
            studentGrades[courseName] = 'PASS' if int(grade) >= PASS_MARK else 'FAIL'

        assessmentResults[studentName] = {
            'grades': studentGrades,
            'cloudCert': None,
            'optionCert': None,
            'overallScore': None
        }
    return assessmentResults


'''
Reads data from Industry_Cert file with provided path and updates corresponding 'cloudCert' and 'optionCert' values in assessmentResults

returns assessmentResults dict
'''
def readIndustryCerts(filePath, assessmentResults):
    rows = []

    try:
        with open(filePath, 'r') as file:
            reader = csv.reader(file)
            for line in reader:
                rows.append(line)
    except FileNotFoundError:
        print(f"Error opening {filePath}")
        raise

    # Read rows from the file
    # Name is the first col, cloudCert date is the second, optionCert is a one of the following columns
    for row in rows[2:]:
        name = row[0]
        studentResults = assessmentResults[name]
        cloudCert = 'PASS' if hasPassedCert(row[1]) else 'FAIL'
        optionCert = 'FAIL'
        for col in row[2:]:
            if hasPassedCert(col):
                optionCert = 'PASS'
                break

        studentResults['cloudCert'] = cloudCert
        studentResults['optionCert'] = optionCert
    return assessmentResults


'''
Method loops through all provided assessment scores for every student in assessmentResults dict and uses their values to calculate overallScore 
'''
def calculateOverallScore(assessmentResults):
    for studentName in assessmentResults:
        studentGrades = assessmentResults[studentName]
        # Overall scores is 'PASS' only when every assessment result is 'PASS'
        if 'FAIL' in studentGrades['grades'].values() or studentGrades['cloudCert'] == 'FAIL' or studentGrades['optionCert'] == 'FAIL':
            studentGrades['overallScore'] = 'FAIL'
        else:
            studentGrades['overallScore'] = 'PASS'
    return assessmentResults


'''
Method writes student results data into a dedicated xlsx file in the output folder
'''
def generateXlslFile(studentResults, name, outputFolder):
    # Does not generate report if overallScore was not calculated
    if studentResults['overallScore'] not in ['PASS', 'FAIL']:
        print(f"Warning: overall score not calculated for {name}, does not generate report.")
        return
    
    generationDate = date.today().strftime("%d/%m/%Y")

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Summary of Results"
    ws["A2"] = "Student Name"
    ws["B2"] = name
    ws["A3"] = "Date"
    ws["B3"] = generationDate
    ws["A5"] = "Assessment Results"

    row = 6
    # Using for loop to write grades for every basic course
    for courseName, grade in studentResults['grades'].items():
        ws.cell(row=row, column=1, value=courseName)
        ws.cell(row=row, column=2, value=grade)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Industry Certification Results")
    row += 1
    ws.cell(row=row, column=1, value="Introduction to Cloud Development")
    ws.cell(row=row, column=2, value=studentResults['cloudCert'])
    row += 1
    ws.cell(row=row, column=1, value="Option Certification")
    ws.cell(row=row, column=2, value=studentResults['optionCert'])
    row += 2
    ws.cell(row=row, column=1, value="Overall Result")
    ws.cell(row=row, column=2, value=studentResults["overallScore"])

    fileName = f"{outputFolder}/{name} - Summary of Results.xlsx"
    # This block prevent script from crashing if file cannot be saved
    try:
        wb.save(fileName)
    except PermissionError:
        print(f"Error: cannot save {fileName} - please close the file.")
        return

    if os.path.exists(fileName):
        print(f"Generated: {fileName}")
    else:
        print(f"Error: file was not saved: {fileName}")


'''
Method formats student results data in a readable way, adds data into an output file and saves the file in the output folder
'''
def generateTxtResult(studentResults, name, outputFolder):
    generationDate = date.today().strftime("%d/%m/%Y")

    lines = []
    lines.append("=" * 90)
    lines.append("Summary Results \n")
    lines.append(f"{'Student Name':<65} | {name}")
    lines.append(f"{'Date Generated':<65} | {generationDate}")
    lines.append("=" * 90)
    lines.append("Assessment Results")
    lines.append("-" * 90)

    for moduleName, result in studentResults['grades'].items():
        lines.append(f"{moduleName:<65} | {result}")

    lines.append("=" * 90)
    lines.append("Industry Certification Results")
    lines.append("-" * 90)
    lines.append(f"{'Introduction to Cloud Development':<65} | {studentResults['cloudCert']}")
    lines.append(f"{'Option Certification':<65} | {studentResults['optionCert']}")
    lines.append("=" * 90)
    lines.append(f"{'Overall Result':<65} | {studentResults['overallScore']}")
    lines.append("=" * 90)

    # Create a file add write lines list into it
    fileName = f"{outputFolder}/{name} - Summary of Results.txt"
    with open(fileName, "w") as f:
        f.write("\n".join(lines))

    print(f"Generated: {fileName}")


def hasPassedCert(certResult):
    # Check if passed certResult string is success of failure
    if certResult == 'FAIL' or certResult == '':
        return False
    return True

# Launch program execution
main()
